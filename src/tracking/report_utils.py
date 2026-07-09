"""
report_utils.py

Shared logic for building the grouped, merged-cell tasking-style
report layout, used by both grouped_report.py (single sheet, color
coded by confidence) and confidence_tabs_report.py (one sheet per
confidence category, no color coding).

Keeping this in one place means both scripts stay consistent --
same Design Point numbering, same grouping rule (exact Date + Time
(Zulu) match), same column layout -- instead of two copies of the
same logic drifting apart over time.
"""

import os

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Imported lazily to avoid circular imports when report_utils is loaded
# before historical_accuracy.py has defined BEHAVIOR_STYLE.
_BEHAVIOR_STYLE_CACHE = None

def _get_behavior_style():
    global _BEHAVIOR_STYLE_CACHE
    if _BEHAVIOR_STYLE_CACHE is None:
        try:
            from historical_accuracy import BEHAVIOR_STYLE
            _BEHAVIOR_STYLE_CACHE = BEHAVIOR_STYLE
        except Exception:
            _BEHAVIOR_STYLE_CACHE = {}
    return _BEHAVIOR_STYLE_CACHE
from openpyxl.utils import get_column_letter

HEADERS = ["Date", "Design Point", "Time (Zulu)", "Target Name", "Target Orbit", "Target NORAD", "Orbit Behavior", "Launch Date", "Country", "Object Type"]
COLUMN_WIDTHS = [14, 13, 12, 22, 13, 14, 18, 13, 9, 14]

# Fill colors (hex, no '#') and matching text colors per confidence
# category. "Not evaluated" covers satellites present in the
# visibility results but missing from the historical accuracy
# report (e.g. if HISTORICAL_MAX_SATELLITES capped that run).
CONFIDENCE_STYLE = {
    "High":                          {"fill": "C6EFCE", "font": "006100"},
    "Moderate":                      {"fill": "FFEB9C", "font": "9C6500"},
    "Low":                           {"fill": "FFD966", "font": "7F6000"},
    "Very low":                      {"fill": "FFC7CE", "font": "9C0006"},
    "Insufficient historical data":  {"fill": "D9D9D9", "font": "404040"},
    "Query failed":                  {"fill": "E4DFEC", "font": "5C4B82"},
    "Not evaluated":                 {"fill": "FFFFFF", "font": "000000"},
}

# Display/priority order for confidence categories wherever they're
# listed or split into tabs.
CONFIDENCE_ORDER = [
    "High", "Moderate", "Low", "Very low",
    "Insufficient historical data", "Query failed", "Not evaluated",
]

GROUP_HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
GROUP_HEADER_FONT = Font(bold=True)
GROUP_DIVIDER = Side(style="thin", color="999999")

PLAIN_FONT = Font(color="000000")


# Confidence category, underlying confidence-score range (see
# accuracy_model.py's thresholds), and what it means. Shared so the
# legend in grouped_report.py and confidence_tabs_report.py always
# says the exact same thing.
CONFIDENCE_LEGEND = [
    ("High", "80-100",
     "Orbit historically very stable -- high confidence in this prediction."),
    ("Moderate", "50-79",
     "Some historical variability observed -- moderate confidence."),
    ("Low", "20-49",
     "Notable historical instability (drag decay, maneuvers) -- treat with caution."),
    ("Very low", "0-19",
     "Highly unstable historical orbit -- low confidence in this prediction."),
    ("Insufficient historical data", "N/A",
     "No/minimal track record (e.g. recently launched) -- model-only estimate, not history-backed. Not scored on the 0-100 scale since no stability could be measured either way."),
    ("Query failed", "N/A",
     "Historical lookup failed/timed out even after retry -- re-run historical_accuracy.py to retry just this satellite. Not a finding about the satellite itself."),
    ("Not evaluated", "N/A",
     "Not included in the historical accuracy evaluation run (e.g. HISTORICAL_MAX_SATELLITES cap)."),
]


def write_legend_sheet(wb, apply_color=True, sheet_name="Legend"):
    """
    Add a Legend sheet to workbook wb listing each confidence
    category, its underlying confidence-score range (0-100), and
    what it means.

    apply_color: if True, the Confidence Category cell is filled
    with that category's color (grouped_report.py's usage, since
    that file's data rows are also color coded). If False, the
    legend is plain text with no fill (confidence_tabs_report.py's
    usage, since that file has no color coding elsewhere -- a
    colored legend there would misleadingly imply colors exist in
    the data tabs).
    """

    ws = wb.create_sheet(sheet_name)

    headers = ["Confidence Category", "Confidence Score Range", "Meaning"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for i, (cat, score_range, desc) in enumerate(CONFIDENCE_LEGEND, start=2):

        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=score_range)
        ws.cell(row=i, column=3, value=desc)

        if apply_color:
            style = CONFIDENCE_STYLE[cat]
            fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
            font = Font(color=style["font"])
            ws.cell(row=i, column=1).fill = fill
            ws.cell(row=i, column=1).font = font

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 80

    for row in ws.iter_rows(min_row=2, max_row=len(CONFIDENCE_LEGEND) + 1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    return ws


def load_merged_data(output_file, accuracy_file):
    """
    Load visible_satellites.xlsx + historical_accuracy_report.xlsx,
    join them on NORAD ID, sort chronologically, and compute the
    Design Point number for each distinct exact (Date, Time (Zulu))
    group.

    Returns (merged_df, design_point_lookup), where design_point_lookup
    maps (date_str, time_str) -> design point number (1, 2, 3...) in
    chronological order. Both report scripts use the SAME lookup so
    Design Point numbers mean the same thing everywhere, regardless
    of which subset of rows ends up on a given sheet.
    """

    if not os.path.exists(output_file):
        raise RuntimeError(f"{output_file} not found. Run main.py first.")

    if not os.path.exists(accuracy_file):
        raise RuntimeError(
            f"{accuracy_file} not found. Run historical_accuracy.py first."
        )

    vis_df = pd.read_excel(output_file, dtype={"Time (Zulu)": str})
    acc_df = pd.read_excel(accuracy_file)

    acc_lookup = (
        acc_df[["Target NORAD", "Confidence Category"]]
        .drop_duplicates(subset="Target NORAD", keep="last")
        .set_index("Target NORAD")["Confidence Category"]
        .to_dict()
    )

    # Build a lookup for both confidence category and orbit behavior.
    # Normalise NORAD IDs to int in both the lookup and the merge so
    # type mismatches (int vs float vs string) never silently produce
    # "Unknown" for every satellite.
    if "Orbit Behavior" in acc_df.columns and acc_df["Orbit Behavior"].notna().any():
        acc_behavior_lookup = (
            acc_df
            .assign(_norad=acc_df["Target NORAD"].astype(int))
            .drop_duplicates(subset="_norad", keep="last")
            .set_index("_norad")["Orbit Behavior"]
            .to_dict()
        )
        _behavior_source = "file"
    else:
        acc_behavior_lookup = {}
        _behavior_source = "missing"

    merged = vis_df.copy()
    merged["Time (Zulu)"] = merged["Time (Zulu)"].astype(str).str.zfill(4)
    merged["Confidence Category"] = (
        merged["Target NORAD"].astype(int).map(acc_lookup).fillna("Not evaluated")
    )
    merged["Orbit Behavior"] = (
        merged["Target NORAD"].astype(int).map(acc_behavior_lookup)
        .fillna("Not evaluated" if _behavior_source == "missing" else "Unknown")
    )
    if _behavior_source == "missing":
        import warnings
        warnings.warn(
            "Accuracy file has no Orbit Behavior data. "
            "Re-run historical_accuracy.py to add behavior detection.",
            stacklevel=2,
        )

    # Join launch metadata columns from the accuracy file.
    # These columns are present when historical_accuracy.py has been
    # run with the current version that fetches SATCAT data.
    _LAUNCH_COLS = {
        "Launch Date":     "launch_date_lkp",
        "Intl Designator": "intl_des_lkp",
        "Object Type":     "obj_type_lkp",
        "Country":         "country_lkp",
        "Launch Site":     "launch_site_lkp",
        "Size Class":      "size_class_lkp",
        "Decay Date":      "decay_date_lkp",
        "GCAT Owner":      "gcat_owner_lkp",
        "GCAT Status":     "gcat_status_lkp",
    }
    for col, tmp in _LAUNCH_COLS.items():
        if col in acc_df.columns:
            lkp = (
                acc_df
                .assign(_n=acc_df["Target NORAD"].astype(int))
                .drop_duplicates(subset="_n", keep="last")
                .set_index("_n")[col]
                .to_dict()
            )
            merged[col] = merged["Target NORAD"].astype(int).map(lkp).fillna("")
        else:
            merged[col] = ""

    merged["_sort_dt"] = pd.to_datetime(
        merged["Date"] + " " + merged["Time (Zulu)"],
        format="%d-%b-%Y %H%M"
    )
    merged = merged.sort_values(by=["_sort_dt", "Target Orbit", "Target Name"])

    group_keys_in_order = list(
        dict.fromkeys(zip(merged["Date"], merged["Time (Zulu)"]))
    )
    design_point_lookup = {key: i + 1 for i, key in enumerate(group_keys_in_order)}

    return merged, design_point_lookup


def write_grouped_sheet(ws, df_subset, design_point_lookup, apply_color=True, start_row=None):
    """
    Write df_subset (already sorted chronologically) into worksheet
    ws using the grouped/merged-cell layout: Date / Design Point /
    Time (Zulu) merged across each group's row span, with a thin
    divider above each new group.

    apply_color: if True, Target Name/Orbit/NORAD columns are filled
    by confidence category (grouped_report's behavior). If False,
    those columns get plain formatting with no fill (used when each
    sheet already represents a single confidence category, so the
    color would be redundant).

    start_row: if provided, the header row is written at this row
    number instead of row 1. Used by generate_tabs_report() to
    place data below the per-tab description block. Defaults to
    None which means row 1 (normal behaviour).

    Design Point numbers come from the shared design_point_lookup
    (built once, across the FULL dataset) so they stay consistent
    with other sheets/files even when df_subset is a filtered
    portion of the data.
    """
    if start_row is None:
        start_row = 1

    # Write the column header row at start_row
    for col, header in enumerate(HEADERS, start=1):
        c = ws.cell(row=start_row, column=col, value=header)
        c.font = Font(bold=True)
    current_row = start_row + 1

    last_group_key = None
    group_spans = {}

    for _, row in df_subset.iterrows():

        group_key = (row["Date"], row["Time (Zulu)"])
        is_first_in_group = (group_key != last_group_key)
        last_group_key = group_key

        # Write row cells directly so we control the exact row
        # number when start_row > 1 (tabs report header block).
        row_values = [
            row["Date"] if is_first_in_group else "",
            design_point_lookup[group_key] if is_first_in_group else "",
            row["Time (Zulu)"] if is_first_in_group else "",
            row["Target Name"],
            row["Target Orbit"],
            row["Target NORAD"],
            row.get("Orbit Behavior", ""),
            row.get("Launch Date", ""),
            row.get("Country", ""),
            row.get("Object Type", ""),
        ]
        for col, val in enumerate(row_values, start=1):
            ws.cell(row=current_row, column=col, value=val)

        excel_row = current_row
        current_row += 1

        if is_first_in_group:
            group_spans[group_key] = [excel_row, excel_row]
        else:
            group_spans[group_key][1] = excel_row

        for col in (1, 2, 3):
            cell = ws.cell(row=excel_row, column=col)
            cell.fill = GROUP_HEADER_FILL
            cell.font = GROUP_HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if apply_color:
            style = CONFIDENCE_STYLE.get(
                row["Confidence Category"], CONFIDENCE_STYLE["Not evaluated"]
            )
            fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
            font = Font(color=style["font"])
        else:
            fill = None
            font = PLAIN_FONT

        for col in (4, 5, 6):
            cell = ws.cell(row=excel_row, column=col)
            if fill is not None:
                cell.fill = fill
            cell.font = font

        # Column 7: Orbit Behavior -- styled by behavior type
        behavior_val = row.get("Orbit Behavior", "")
        bstyle = _get_behavior_style().get(behavior_val, {"fill": "FFFFFF", "font": "000000"})
        bfill  = PatternFill(start_color=bstyle["fill"], end_color=bstyle["fill"],
                             fill_type="solid")
        bfont  = Font(color=bstyle["font"], size=9)
        bcell  = ws.cell(row=excel_row, column=7)
        bcell.fill = bfill
        bcell.font = bfont
        bcell.alignment = Alignment(horizontal="center", vertical="center")

        # Columns 8-10: Launch Date, Country, Object Type -- plain,
        # centered, slightly smaller font for the metadata fields
        for col in (8, 9, 10):
            mc = ws.cell(row=excel_row, column=col)
            mc.font      = Font(size=9, color="404040")
            mc.alignment = Alignment(horizontal="center", vertical="center")

    for (grp_start, grp_end) in group_spans.values():

        if grp_end > grp_start:
            for col in (1, 2, 3):
                ws.merge_cells(
                    start_row=grp_start, start_column=col,
                    end_row=grp_end, end_column=col
                )

        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=grp_start, column=col)
            cell.border = Border(top=GROUP_DIVIDER)
            if col == 7:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = f"A{start_row + 1}"

    return len(df_subset), len(group_spans)
