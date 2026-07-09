"""
reports.py

Generates the two Excel report files from the visibility results
and historical confidence scores. Called automatically by main.py
(Step 8) and can also be run standalone.

  generate_grouped_report()  -- one sheet, color-coded by confidence
  generate_tabs_report()     -- one tab per confidence category, with
                                an Overview sheet and per-tab headers

Run standalone (uses last_run.json to find the right files):
    python reports.py
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_FILE
from report_utils import (
    load_merged_data, write_grouped_sheet, write_legend_sheet,
    CONFIDENCE_ORDER, CONFIDENCE_LEGEND, CONFIDENCE_STYLE,
)

# Behavior style imported lazily (defined in historical_accuracy.py)
def _behavior_style():
    try:
        from historical_accuracy import BEHAVIOR_STYLE
        return BEHAVIOR_STYLE
    except Exception:
        return {}
from run_state import get_run_files


# ── Per-tab descriptions ───────────────────────────────────────
# Each entry: (score_range, short_headline, detailed_explanation)
_TAB_DESCRIPTIONS = {
    "High": (
        "Score 80-100",
        "Orbit historically very stable — predictions are reliable.",
        "Satellites on this tab have shown minimal variation in altitude "
        "and inclination over the historical lookback window. Their future "
        "positions can be predicted with high confidence. Little to no "
        "additional scrutiny is required before tasking."
    ),
    "Moderate": (
        "Score 50-79",
        "Some historical variability observed — predictions are probably reliable.",
        "Satellites on this tab show minor orbital variability (e.g. small "
        "amounts of atmospheric drag or occasional station-keeping). "
        "Predictions are generally reliable but worth a secondary check "
        "for high-priority tasking windows."
    ),
    "Low": (
        "Score 20-49",
        "Notable instability detected — treat predictions with caution.",
        "Satellites on this tab have a history of meaningful orbital changes "
        "such as drag decay, altitude adjustments, or maneuvers. Predicted "
        "pass times and azimuths may differ noticeably from actual. Increase "
        "search dwell time or widen the acquisition window when tasking."
    ),
    "Very low": (
        "Score 0-19",
        "Highly unstable orbit — predictions may be significantly wrong.",
        "Satellites on this tab have shown high orbital instability. "
        "Confidence in predicted positions is low. These satellites should "
        "be treated as best-effort only — the actual pass may differ "
        "substantially in time, azimuth, and elevation from the prediction."
    ),
    "Insufficient historical data": (
        "Score N/A",
        "No historical track record — model-only estimate.",
        "Satellites on this tab have fewer than 2 historical TLE records in "
        "the lookback window, typically because they were recently launched. "
        "There is no stability history to draw on. Treat predictions as "
        "model-based extrapolations, not history-backed forecasts."
    ),
    "Query failed": (
        "Score N/A",
        "Space-Track lookup failed — confidence unknown.",
        "Satellites on this tab could not be evaluated because the "
        "Space-Track query timed out or failed even after retry. This is "
        "a transient network issue, not a finding about the satellite itself. "
        "Re-run historical_accuracy.py to retry these satellites."
    ),
    "Not evaluated": (
        "Score N/A",
        "Not included in the confidence evaluation run.",
        "Satellites on this tab were visible in the analysis window but "
        "were not included in the historical accuracy evaluation run "
        "(e.g. the HISTORICAL_MAX_SATELLITES cap was reached). Re-run "
        "historical_accuracy.py with the cap removed to score these."
    ),
}


# ── Styling constants ──────────────────────────────────────────
_TITLE_FONT      = Font(name="Calibri", size=14, bold=True, color="1F3864")
_HEADING_FONT    = Font(name="Calibri", size=11, bold=True, color="1F3864")
_LABEL_FONT      = Font(name="Calibri", size=10, bold=True, color="404040")
_VALUE_FONT      = Font(name="Calibri", size=10, color="000000")
_DESC_FONT       = Font(name="Calibri", size=10, italic=True, color="404040")
_HEADER_FILL     = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_HEADER_FONT_W   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_META_FILL       = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
_SECTION_FILL    = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_THIN_BORDER     = Border(
    bottom=Side(style="thin", color="BFBFBF")
)


def _resolve_files(output_file, accuracy_file, report_file, report_key):
    """Resolve file paths from arguments or last_run.json."""
    if output_file is None or accuracy_file is None or report_file is None:
        run = get_run_files()
        output_file   = output_file   or run["visible_satellites"]
        accuracy_file = accuracy_file or run["accuracy_report"]
        report_file   = report_file   or run[report_key]
    return output_file, accuracy_file, report_file


def _get_run_info(output_file):
    """
    Pull analysis metadata from last_run.json if available,
    otherwise extract what we can from the output filename.
    Returns a dict with date, local_window, zulu_window, sensor,
    generated_at, and tag.
    """
    run = get_run_files()
    raw = run.get("_raw", {})

    # Prefer stored metadata from last_run.json
    if raw.get("date"):
        date_str    = raw.get("date", "")
        tz_label    = raw.get("timezone_label", "Zulu")
        local_start = raw.get("local_start_hhmm", "")
        local_end   = raw.get("local_end_hhmm", "")
        utc_start   = raw.get("utc_start_hhmm", "")
        utc_end     = raw.get("utc_end_hhmm", "")
        tag         = raw.get("tag", "")

        local_window = (
            f"{local_start}-{local_end} {tz_label}" if local_start else "N/A"
        )
        zulu_window  = f"{utc_start}-{utc_end}Z" if utc_start else "N/A"
    else:
        # Fall back: parse tag from filename e.g. 2026-08-05_1100-2300Z
        tag = run.get("tag", "")
        parts = tag.split("_")
        date_str    = parts[0] if parts else "Unknown"
        zulu_window = parts[1] if len(parts) > 1 else "Unknown"
        local_window = "See Zulu time"

    return {
        "date":         date_str,
        "local_window": local_window,
        "zulu_window":  zulu_window,
        "tag":          tag,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _write_overview_sheet(wb, run_info, merged):
    """
    Create an Overview sheet as the first sheet in the workbook.
    Includes:
      - Report title and analysis metadata (date, time window)
      - Satellite count summary by confidence category
      - Brief explanation of each category present in this run
    """
    ws = wb.create_sheet("Overview", 0)   # insert at position 0 = first tab

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    row = 1

    # ── Title block ─────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Satellite Visibility — Final Report With Confidence Tabs")
    ws.cell(row=row, column=1).font      = _TITLE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row, column=1).fill      = _META_FILL
    ws.row_dimensions[row].height = 24
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1,
            value="Each tab in this workbook contains the satellites "
                  "visible during the analysis window, filtered by confidence level.")
    ws.cell(row=row, column=1).font      = _DESC_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center",
                                                     wrap_text=True)
    ws.cell(row=row, column=1).fill      = _META_FILL
    ws.row_dimensions[row].height = 28
    row += 2

    # ── Analysis metadata ────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Analysis Details")
    ws.cell(row=row, column=1).font = _HEADING_FONT
    ws.cell(row=row, column=1).fill = _SECTION_FILL
    ws.cell(row=row, column=1).border = _THIN_BORDER
    row += 1

    meta = [
        ("Analysis Date",          run_info["date"]),
        ("Local Time Window",      run_info["local_window"]),
        ("Zulu (UTC) Time Window", run_info["zulu_window"]),
        ("Report Generated",       run_info["generated_at"]),
        ("Run Tag",                run_info["tag"]),
    ]
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = _VALUE_FONT
        ws.cell(row=row, column=1).fill = _META_FILL
        ws.cell(row=row, column=2).fill = _META_FILL
        row += 1

    row += 1

    # ── Satellite count summary ──────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Satellite Count by Confidence Category")
    ws.cell(row=row, column=1).font = _HEADING_FONT
    ws.cell(row=row, column=1).fill = _SECTION_FILL
    ws.cell(row=row, column=1).border = _THIN_BORDER
    row += 1

    # Table header
    for col, header in enumerate(["Confidence Category", "Score Range",
                                   "Satellites", "% of Total"], start=1):
        c = ws.cell(row=row, column=col, value=header)
        c.font   = _HEADER_FONT_W
        c.fill   = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1

    total_sats = len(merged)
    for category in CONFIDENCE_ORDER:
        subset = merged[merged["Confidence Category"] == category]
        count  = len(subset)
        if count == 0:
            continue

        # Score range from legend
        score_range = next(
            (sr for cat, sr, _ in CONFIDENCE_LEGEND if cat == category), "N/A"
        )
        pct = f"{count / total_sats * 100:.1f}%" if total_sats else "0%"

        style = CONFIDENCE_STYLE.get(category, {"fill": "FFFFFF", "font": "000000"})
        fill  = PatternFill(start_color=style["fill"],
                            end_color=style["fill"], fill_type="solid")
        font  = Font(color=style["font"])

        ws.cell(row=row, column=1, value=category).fill = fill
        ws.cell(row=row, column=1).font      = font
        ws.cell(row=row, column=2, value=score_range).fill = fill
        ws.cell(row=row, column=2).font      = font
        ws.cell(row=row, column=3, value=count).fill = fill
        ws.cell(row=row, column=3).font      = font
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=pct).fill = fill
        ws.cell(row=row, column=4).font      = font
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = _LABEL_FONT
    ws.cell(row=row, column=3, value=total_sats).font = _LABEL_FONT
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=4, value="100%").font = _LABEL_FONT
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    row += 2

    # ── Tab guide ────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Tab Guide — What Each Tab Contains")
    ws.cell(row=row, column=1).font = _HEADING_FONT
    ws.cell(row=row, column=1).fill = _SECTION_FILL
    ws.cell(row=row, column=1).border = _THIN_BORDER
    row += 1

    # Only describe tabs that actually appear in this workbook
    present_categories = [
        cat for cat in CONFIDENCE_ORDER
        if len(merged[merged["Confidence Category"] == cat]) > 0
    ]

    for category in present_categories:
        subset = merged[merged["Confidence Category"] == category]
        count  = len(subset)
        desc   = _TAB_DESCRIPTIONS.get(category, ("N/A", category, ""))
        score_range, headline, detail = desc

        style = CONFIDENCE_STYLE.get(category, {"fill": "FFFFFF", "font": "000000"})
        fill  = PatternFill(start_color=style["fill"],
                            end_color=style["fill"], fill_type="solid")
        cfont = Font(color=style["font"], bold=True)

        # Category name cell (colored)
        ws.merge_cells(f"A{row}:A{row+2}")
        c = ws.cell(row=row, column=1, value=f"{category}\n({count} satellites)")
        c.font      = cfont
        c.fill      = fill
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)

        # Headline
        ws.merge_cells(f"B{row}:D{row}")
        ws.cell(row=row, column=2, value=f"{score_range}  —  {headline}")
        ws.cell(row=row, column=2).font      = _LABEL_FONT
        ws.cell(row=row, column=2).alignment = Alignment(vertical="top")

        # Detail
        ws.merge_cells(f"B{row+1}:D{row+2}")
        ws.cell(row=row+1, column=2, value=detail)
        ws.cell(row=row+1, column=2).font      = _DESC_FONT
        ws.cell(row=row+1, column=2).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        ws.row_dimensions[row+1].height = 40

        row += 3

    ws.freeze_panes = "A1"
    return ws


def _write_tab_header(ws, category, run_info, sat_count):
    """
    Insert a 4-row description block at the top of a confidence tab
    before the data rows, explaining what the tab is and what the
    data covers.

    Returns the row number where the data header should start.
    """
    desc        = _TAB_DESCRIPTIONS.get(category, ("N/A", category, ""))
    score_range, headline, detail = desc

    style = CONFIDENCE_STYLE.get(category, {"fill": "F2F2F2", "font": "000000"})
    fill  = PatternFill(start_color=style["fill"],
                        end_color=style["fill"], fill_type="solid")
    cfont = Font(color=style["font"], bold=True, size=12)

    # Row 1: Category title
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1,
                value=f"{category}  ({score_range})  —  {sat_count} satellites")
    c.font      = cfont
    c.fill      = fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    # Row 2: Headline
    ws.merge_cells("A2:F2")
    ws.cell(row=2, column=1, value=headline)
    ws.cell(row=2, column=1).font      = _LABEL_FONT
    ws.cell(row=2, column=1).fill      = _META_FILL
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Row 3: Detail explanation
    ws.merge_cells("A3:F3")
    ws.cell(row=3, column=1, value=detail)
    ws.cell(row=3, column=1).font      = _DESC_FONT
    ws.cell(row=3, column=1).fill      = _META_FILL
    ws.cell(row=3, column=1).alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )
    ws.row_dimensions[3].height = 42

    # Row 4: Analysis window line
    ws.merge_cells("A4:F4")
    ws.cell(row=4, column=1,
            value=f"Analysis date: {run_info['date']}  |  "
                  f"Zulu window: {run_info['zulu_window']}  |  "
                  f"Local window: {run_info['local_window']}")
    ws.cell(row=4, column=1).font      = _DESC_FONT
    ws.cell(row=4, column=1).fill      = _META_FILL
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 16

    # Row 5: blank spacer
    ws.row_dimensions[5].height = 6

    # Data starts at row 6
    return 6


def generate_grouped_report(output_file=None, accuracy_file=None, report_file=None):
    """
    Cross-reference visibility results with confidence scores and
    produce a single color-coded grouped tasking-sheet Excel file.

    Satellites are grouped by exact (Date, Time Zulu) and given
    Design Point numbers in chronological order. Each satellite row
    is color-coded by confidence level (green=High through red=Very
    low). A Legend sheet explains the color coding and score ranges.
    """
    output_file, accuracy_file, report_file = _resolve_files(
        output_file, accuracy_file, report_file, "grouped_report"
    )

    merged, dp_lookup = load_merged_data(output_file, accuracy_file)

    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Report"

    num_rows, num_groups = write_grouped_sheet(ws, merged, dp_lookup, apply_color=True)
    write_legend_sheet(wb, apply_color=True)

    os.makedirs(os.path.dirname(report_file), exist_ok=True)
    wb.save(report_file)
    print(f"Grouped report: {num_rows} satellites across {num_groups} time groups")
    print(f"  → {report_file}")


def _write_behavior_legend(wb):
    """Add a Behavior Legend sheet explaining the orbit behavior flags."""
    from openpyxl.styles import PatternFill, Font, Alignment
    ws = wb.create_sheet("Behavior Legend")

    bstyle = _behavior_style()

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70

    # Header
    ws.cell(row=1, column=1, value="Orbit Behavior Flag").font = Font(bold=True)
    ws.cell(row=1, column=2, value="What it means").font = Font(bold=True)

    behavior_rows = [
        ("Stable",           "No significant orbital changes detected. Predictions are as reliable as the confidence score indicates."),
        ("Decaying",         "Normal atmospheric drag decay for an uncontrolled LEO object. Altitude is declining slowly at a predictable rate."),
        ("Rapid Decay",      "Altitude declining fast (> 0.5 km/day). Object may reenter within months. Use the most recent TLE available."),
        ("Maneuvering",      "1-2 maneuvers detected in the last 90 days. Prediction accuracy is reduced until the orbit stabilises post-burn."),
        ("Recently Changed", "Orbit has settled at a new altitude compared to earlier history, but no maneuvers detected in last 90 days."),
        ("Actively Managed", "3+ maneuvers in last 90 days. Orbit may change at any time. Treat predictions as best-effort only."),
        ("Insufficient Data","Too few historical records to classify behavior (e.g. recently launched)."),
    ]

    for i, (flag, desc) in enumerate(behavior_rows, start=2):
        style = bstyle.get(flag, {"fill": "FFFFFF", "font": "000000"})
        fill  = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
        font  = Font(color=style["font"])
        ws.cell(row=i, column=1, value=flag).fill = fill
        ws.cell(row=i, column=1).font = font
        ws.cell(row=i, column=2, value=desc)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 28

    return ws


def generate_tabs_report(output_file=None, accuracy_file=None, report_file=None):
    """
    Cross-reference visibility results with confidence scores and
    produce an Excel file with:
      - An Overview sheet (first tab): analysis date/time, satellite
        count summary by confidence level, and a guide to each tab
      - One tab per confidence category: satellites in that category
        with a description header showing what the tab covers
      - A Legend sheet (last tab): score ranges and meanings

    Only categories with at least one satellite are included.
    """
    output_file, accuracy_file, report_file = _resolve_files(
        output_file, accuracy_file, report_file, "tabs_report"
    )

    merged, dp_lookup = load_merged_data(output_file, accuracy_file)
    run_info = _get_run_info(output_file)

    wb = Workbook()
    wb.remove(wb.active)   # drop the default blank sheet

    # Overview sheet first
    _write_overview_sheet(wb, run_info, merged)

    total_rows = 0
    sheets     = 0

    for category in CONFIDENCE_ORDER:
        subset = merged[merged["Confidence Category"] == category]
        if len(subset) == 0:
            continue

        ws = wb.create_sheet(category)

        # Write the description header block, get the row where data starts
        data_start_row = _write_tab_header(ws, category, run_info, len(subset))

        # Write the satellite data starting below the header
        num_rows, _ = write_grouped_sheet(
            ws, subset, dp_lookup,
            apply_color=False,
            start_row=data_start_row,
        )

        total_rows += num_rows
        sheets     += 1
        print(f"  {category}: {num_rows} satellites")

    if sheets == 0:
        print("No satellites found in any confidence category -- nothing to save.")
        return

    _write_behavior_legend(wb)
    write_legend_sheet(wb, apply_color=False)

    os.makedirs(os.path.dirname(report_file), exist_ok=True)
    wb.save(report_file)
    print(f"Final report: {total_rows} satellites across {sheets} data tabs")
    print(f"  → {report_file}")


if __name__ == "__main__":
    print("Generating reports from last run...")
    generate_grouped_report()
    print()
    generate_tabs_report()
